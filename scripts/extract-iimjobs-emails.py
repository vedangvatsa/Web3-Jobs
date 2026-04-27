#!/usr/bin/env python3
"""
Extract emails from iimjobs.xlsx and filter to keep only 'safe' personal emails.
Removes company/edu/org/college/university domains that are likely dead.
"""

import openpyxl
import re
import json
from collections import Counter

INPUT_FILE = "/Users/vedang/Downloads/iimjobs.xlsx"
OUTPUT_DIR = "LOCAL_PATH/scripts"

# Major free email providers that are still active and deliverable
SAFE_DOMAINS = {
    # Google
    "gmail.com", "googlemail.com",
    # Microsoft
    "outlook.com", "hotmail.com", "live.com", "msn.com",
    "outlook.in", "hotmail.co.in", "live.in",
    # Yahoo
    "yahoo.com", "yahoo.co.in", "yahoo.in", "ymail.com",
    "yahoo.co.uk", "yahoo.ca", "yahoo.com.au",
    # Apple
    "icloud.com", "me.com", "mac.com",
    # Others
    "protonmail.com", "proton.me",
    "zoho.com", "zohomail.com",
    "aol.com",
    "mail.com",
    "gmx.com", "gmx.net",
    "rediffmail.com", "rediff.com",
    "inbox.com",
    "fastmail.com",
    "tutanota.com", "tuta.io",
    "yandex.com", "yandex.ru",
    "mail.ru",
}

# Patterns that indicate risky/dead domains
RISKY_PATTERNS = [
    r"\.edu$", r"\.edu\.", r"\.ac\.", r"\.ac$",
    r"\.gov$", r"\.gov\.",
    r"\.mil$",
    r"\.org$",   # nonprofits often rotate emails
    r"\.college$", r"\.university$",
    r"\.nic\.in$",  # Indian government
]

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")

def main():
    print(f"📂 Reading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]

    all_emails = []
    safe_emails = set()
    unsafe_emails = set()
    invalid_emails = []
    domain_counts = Counter()
    
    header_skipped = False
    row_count = 0
    
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        
        row_count += 1
        if row_count % 100000 == 0:
            print(f"  ... processed {row_count:,} rows")
        
        # Email is in column 3 (index 2)
        email = row[2] if len(row) > 2 else None
        if not email or not isinstance(email, str):
            continue
        
        email = email.strip().lower()
        
        # Basic validation
        if not EMAIL_RE.match(email):
            invalid_emails.append(email)
            continue
        
        domain = email.split("@")[1]
        domain_counts[domain] += 1
        
        # Check if it's a known safe personal provider
        if domain in SAFE_DOMAINS:
            safe_emails.add(email)
        else:
            # Check for risky patterns
            is_risky = any(re.search(p, domain) for p in RISKY_PATTERNS)
            if is_risky:
                unsafe_emails.add(email)
            else:
                # Unknown domain — likely a company email, risky
                unsafe_emails.add(email)
    
    wb.close()
    
    # Remove duplicates that are already in existing manual-emails files
    existing = set()
    for f in ["manual-emails.json", "manual-emails-2.json", "manual-emails-3.json", "manual-emails-4.json"]:
        try:
            with open(f"{OUTPUT_DIR}/{f}") as fh:
                existing.update(e.lower().strip() for e in json.load(fh))
        except Exception:
            pass
    
    new_safe = safe_emails - existing
    
    # Stats
    print(f"\n{'='*60}")
    print(f"📊 RESULTS")
    print(f"{'='*60}")
    print(f"Total rows:           {row_count:,}")
    print(f"Valid emails:         {len(safe_emails) + len(unsafe_emails):,}")
    print(f"Invalid/malformed:    {len(invalid_emails):,}")
    print(f"")
    print(f"✅ Safe (personal):   {len(safe_emails):,}")
    print(f"❌ Risky (corp/edu):  {len(unsafe_emails):,}")
    print(f"Already in pipeline:  {len(safe_emails & existing):,}")
    print(f"🆕 New safe emails:   {len(new_safe):,}")
    
    # Top domains
    print(f"\n📈 Top 20 domains:")
    for domain, count in domain_counts.most_common(20):
        marker = "✅" if domain in SAFE_DOMAINS else "❌"
        print(f"  {marker} {domain}: {count:,}")
    
    # Top unsafe domains
    unsafe_domains = {d: c for d, c in domain_counts.items() if d not in SAFE_DOMAINS}
    print(f"\n📈 Top 20 UNSAFE domains (excluded):")
    for domain, count in Counter(unsafe_domains).most_common(20):
        print(f"  ❌ {domain}: {count:,}")
    
    # Save results
    safe_list = sorted(new_safe)
    output_file = f"{OUTPUT_DIR}/iimjobs-safe-emails.json"
    with open(output_file, "w") as f:
        json.dump(safe_list, f, indent=2)
    print(f"\n💾 Saved {len(safe_list):,} new safe emails to {output_file}")
    
    # Also save risky emails for reference
    risky_file = f"{OUTPUT_DIR}/iimjobs-risky-emails.json"
    with open(risky_file, "w") as f:
        json.dump(sorted(unsafe_emails), f, indent=2)
    print(f"💾 Saved {len(unsafe_emails):,} risky emails to {risky_file}")

if __name__ == "__main__":
    main()
